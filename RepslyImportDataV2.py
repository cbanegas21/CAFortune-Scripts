import os
import re
from openpyxl import load_workbook

original_files_directory = "C:\\Users\\carlo\\OneDrive - C.A. Fortune - C.A. Carlin\\Gotham Dashboards\\2024_DataExtracts"

def get_column_index_by_header(ws, header_name):
    for row in ws.iter_rows(max_row=1):
        for cell in row:
            if cell.value == header_name:
                return cell.column
    return None

def extract_url_from_formula(formula):
    url_match = re.search(r'"(https?://[^"]+)"', formula)
    if url_match:
        return url_match.group(1)
    return None

def clean_representative_id(ws):
    # Column A is always 'Representative ID', so we hardcode it
    rep_id_col_idx = 1
    max_row = ws.max_row + 1

    for row in range(2, max_row):
        rep_id_cell = ws.cell(row=row, column=rep_id_col_idx)
        rep_id_value = str(rep_id_cell.value).strip()

        # Check if the value contains any letters or non-numeric characters
        if not rep_id_value.isdigit():
            # Replace the value with 0 if it's not a valid number
            ws.cell(row=row, column=rep_id_col_idx).value = 0

def remove_and_add_columns(file):
    try:
        wb = load_workbook(file)
        ws = wb.active

        extract_col_idx = get_column_index_by_header(ws, 'EXTRACT')
        form_link_col_idx = get_column_index_by_header(ws, 'Form Link')
        link_to_form_col_idx = get_column_index_by_header(ws, 'Link to Form')

        # Remove the columns if they exist
        if form_link_col_idx:
            ws.delete_cols(form_link_col_idx)
        if extract_col_idx:
            ws.delete_cols(extract_col_idx)

        # Re-add the headers in the appropriate columns (assuming extract_col_idx < form_link_col_idx initially)
        max_row = ws.max_row + 1
        ws.insert_cols(extract_col_idx)
        ws.insert_cols(form_link_col_idx)

        ws.cell(row=1, column=extract_col_idx).value = 'EXTRACT'
        ws.cell(row=1, column=form_link_col_idx).value = 'Form Link'

        # Extract the hyperlink from the 'Link to Form' column and populate 'Form Link' column
        for row in range(2, max_row):
            link_cell = ws.cell(row=row, column=link_to_form_col_idx)
            if link_cell.value and isinstance(link_cell.value, str) and link_cell.value.startswith('=HYPERLINK'):
                extracted_url = extract_url_from_formula(link_cell.value)
                ws.cell(row=row, column=form_link_col_idx).value = extracted_url
            else:
                ws.cell(row=row, column=form_link_col_idx).value = None  # If there's no hyperlink, set it to None

            # Add an incremental value to the 'EXTRACT' column
            ws.cell(row=row, column=extract_col_idx).value = row - 1

        # Clean the 'Representative ID' column (Column A)
        clean_representative_id(ws)

        wb.save(file)
        print(f"{file} processed successfully")
    except Exception as e:
        print(f"Error processing {file}: {e}")

def process_all_files(directory):
    for file_name in os.listdir(directory):
        if file_name.endswith(".xlsx"):
            file_path = os.path.join(directory, file_name)
            remove_and_add_columns(file_path)

process_all_files(original_files_directory)

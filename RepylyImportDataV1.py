import os
from openpyxl import load_workbook

# Directorio de archivos originales y directorio de archivos descargados
original_files_directory = "C:\\Users\\carlo\\OneDrive - C.A. Fortune - C.A. Carlin\\Gotham Dashboards\\2024_DataExtracts"
downloaded_files_directory = "C:\\Users\\carlo\\OneDrive - C.A. Fortune - C.A. Carlin\\Gotham Dashboards\\2024_DataExtracts\\Imports for Data Extracts"

# Mapeo de nombres de archivos descargados a nombres de archivos originales
file_mapping = {
    "Beyond Meat - Giant.xlsx": "Beyond Meat - Giant NEW.xlsx",
    "Beyond Meat - Wegmans.xlsx": "Beyond Meat - Wegmans NEW.xlsx",
    "Beyond Meat - Whole Foods.xlsx": "Beyond Meat - Whole Foods NEW.xlsx",
    "Garden of Flavor.xlsx": "Garden of Flavor New.xlsx",
    "GRAZA Whole Foods.xlsx": "GRAZA Whole Foods.xlsx",
    "LaCroix.xlsx": "LaCroix.xlsx",
    "Our Home-All Retail.xlsx": "Our Home All Retail.xlsx",
    "Our Home- NGVC.xlsx": "Our Home NGVC.xlsx",
    "Perfect Snacks Shoprite.xlsx": "Perfect Snacks Shoprite.xlsx",
    "Perfect Snacks Stop & Shop.xlsx": "Perfect Snacks Stop & Shop.xlsx",
    "Saratoga & Poland Springs (BlueTriton).xlsx": "Saratoga and Poland Springs (Blue Triton).xlsx",
    "Terra (Hain).xlsx": "Terra (Hain).xlsx",
    "Vahdam Tea Wegmans.xlsx": "Vahdam Teas.xlsx",
    "Row 7 Nor Cal.xlsx": "Row 7 Nor Cal.xlsx",
    "Row 7 North Atlantic.xlsx": "Row 7 North Atlantic.xlsx",
    "Row 7 Northeast.xlsx": "Row 7 Northeast.xlsx",
    "Row 7 Mid Atlantic.xlsx": "Row 7 Mid Atlantic.xlsx",
    "Row 7 So Cal.xlsx": "Row 7 So Cal.xlsx",
    "Row 7 Pacific Northwest.xlsx": "Row 7 Pacific Northwest.xlsx",
    "Tache.xlsx": "Tache.xlsx",
    "Catalina Crunch Wakefern.xlsx": "Catalina Crunch Wakefern.xlsx",
    "Garden of Eatin Tortilla (Hain).xlsx": "Garden of Eatin Tortilla (Hain) NEW.xlsx",
    "Garden Veggie Straws (Hain).xlsx": "Garden Veggie Straws (Hain) NEW.xlsx",
    "Guayaki Yerbe Mate.xlsx": "Guayaki.xlsx",
    "Hain Independent Celestial Tea.xlsx": "Hain Independent Celestial Tea.xlsx",
    "Pur Gum.xlsx": "Pur Gum NEW.xlsx",
    "Yesly-Jewel.xlsx": "Yesly Jewel.xlsx",
    "Yesly-Stop&Shop.xlsx": "Yesly Stop&Shop.xlsx",
    "Solely-WF.xlsx": "Solely Whole Foods.xlsx",
    "Solely-ShopRite.xlsx": "Solely Shoprite.xlsx",
    "Hal's New York Seltzer.xlsx": "Hals New York Seltzer.xlsx",
    "Beyond Meat - Pavilions.xlsx": "Beyond Meat - Pavilions.xlsx",
    "Beyond Meat - Ralphs.xlsx": "Beyond Meat - Ralphs.xlsx",
    "Beyond Meat - Sprouts.xlsx": "Beyond Meat - Sprouts.xlsx",
    "Beyond Meat - Vons.xlsx": "Beyond Meat - Vons .xlsx",
    "Beyond Meat - Whole Foods - So Cal.xlsx": "Beyond Meat - Whole Foods - So Cal.xlsx",
    "Beyond Meat - Whole Foods.xlsx": "Beyond Meat - Whole Foods NEW.xlsx",
    "Hain Independent Imagine.xlsx": "Hain Independent Imagine.xlsx",
    "Lundberg Natural Grocers.xlsx": "Lundberg Natural Grocers.xlsx",
    "Celestial Tea (Hain).xlsx": "Celestial Tea (Hain).xlsx",
    "Equator Coffee.xlsx": "Equator Coffee.xlsx",
    "St. Dalfour NorCal.xlsx": "St. Dalfour NorCal.xlsx",
    "St. Dalfour NorthEast.xlsx": "St. Dalfour NorthEast.xlsx",
    "FTGU (Our Home) WF.xlsx": "FTGU (Our Home) WF.xlsx",
    "FSTG (Our Home) WF.xlsx": "FSTG (Our Home) WF.xlsx",
    "RW Garcia (Our Home) WF.xlsx": "RW Garcia (Our Home) WF.xlsx",
    "Catalina Crunch Cookie UPC Conversion UNFI.xlsx": "Catalina Crunch Cookie UPC Conversion UNFI.xlsx",
    "Catalina Crunch Cookie UPC Conversion KeHe.xlsx": "Catalina Crunch Cookie UPC Conversion KeHe.xlsx",
    "Maria & Ricardos Bag Drop Audit Sprouts.xlsx": "Maria & Ricardos Bag Drop Audit Sprouts.xlsx",
    "Maria & Ricardos Bag Drop Audit WF.xlsx": "Maria & Ricardos Bag Drop Audit WF.xlsx",
    "Panos Amore Image Project.xlsx": "Panos Amore Image Project.xlsx",
    "Panos Better Than Milk Image Project.xlsx": "Panos Better Than Milk Image Project.xlsx",
    "Panos Chatfields Image Project.xlsx": "Panos Chatfields Image Project.xlsx",
    "Panos Mi-Del Image Project.xlsx": "Panos Mi-Del Image Project.xlsx",
    "Panos Sesmark Image Project.xlsx": "Panos Sesmark Image Project.xlsx",
    "Panos Walden Farms Image Project.xlsx": "Panos Walden Farms Image Project.xlsx",
    "Pop Secret- Wegman's (Our Home).xlsx": "Pop Secret- Wegman's.xlsx",
    "Saratoga Water (Primo Brands) North Atlantic Blitz.xlsx": "Saratoga Water (Primo Brands) North Atlantic Blitz.xlsx",
    "Sola-Whole Foods.xlsx": "Sola-Whole Foods.xlsx",
    "_Panos Ka-Me Image Project.xlsx": "_Panos Ka-Me Image Project.xlsx",
    "Panos.xlsx": "Panos.xlsx",
    "Free2b.xlsx": "Free2b.xlsx",
    "GoodPop Sticker Audit.xlsx": "GoodPop Sticker Audit.xlsx",
    "Sola-Whole Foods.xlsx": "Sola-Whole Foods.xlsx",
    "Lundberg Rice Cakes Natural Independent (Dec-Jan).xlsx": "Lundberg Rice Cakes Natural Independent (Dec-Jan).xlsx",
    "Yesly ShopRite Audit.xlsx": "Yesly ShopRite Audit.xlsx",
    "Back 2 Nature Wakefern.xlsx": "Back 2 Nature Wakefern.xlsx",
    "Beyond Meat Sun Sausage Launch.xlsx": "Beyond Meat Sun Sausage Launch.xlsx",
    "Solely Natural Grocers.xlsx": "Solely Natural Grocers.xlsx",
    "Back 2 Nature Fresh Thyme.xlsx": "Back 2 Nature Fresh Thyme.xlsx",
    "Back 2 Nature Giant Landover.xlsx": "Back 2 Nature Giant Landover.xlsx",
    "Back 2 Nature Hannaford.xlsx": "Back 2 Nature Hannaford.xlsx",
    "Back 2 Nature Harmons.xlsx": "Back 2 Nature Harmons.xlsx",
    "Back 2 Nature INFRA.xlsx": "Back 2 Nature INFRA.xlsx",
    "Back 2 Nature Jewel.xlsx": "Back 2 Nature Jewel.xlsx",
    "Back 2 Nature Meijer.xlsx": "Back 2 Nature Meijer.xlsx",
    "Back 2 Nature MOMs.xlsx": "Back 2 Nature MOMs.xlsx",
    "Back 2 Nature Mother's.xlsx": "Back 2 Nature Mother's.xlsx",
    "Back 2 Nature NCG.xlsx": "Back 2 Nature NCG.xlsx",
    "Back 2 Nature NGVC.xlsx": "Back 2 Nature NGVC .xlsx",
    "Back 2 Nature Sprouts.xlsx": "Back 2 Nature Sprouts.xlsx",
    "Back 2 Nature Stop & Shop.xlsx": "Back 2 Nature Stop & Shop.xlsx",
    "Back 2 Nature Wegmans.xlsx": "Back 2 Nature Wegmans.xlsx",
    "Back 2 Nature Whole Foods.xlsx": "Back 2 Nature Whole Foods.xlsx",
    "Grounded.xlsx": "Grounded.xlsx",
    "Krave Whole Foods.xlsx": "Krave Whole Foods.xlsx",
    "Sayso.xlsx": "Sayso.xlsx",
    "Wilde Whole Foods.xlsx": "Wilde Whole Foods.xlsx",
    "Parm Crisps- Wegman's.xlsx": "Parm Crisps- Wegman's.xlsx",
    "Maria & Ricardos Tortillas Promo.xlsx": "Maria & Ricardos Tortillas Promo.xlsx",
    "Beyond Meat Sprouts Activation (Nation Wide).xlsx": "Beyond Meat Sprouts Activation (Nation Wide).xlsx",
    "Graza Wakefern.xlsx": "Graza Wakefern.xlsx",
    "Lil Gourmets Whole Foods.xlsx": "Lil Gourmets Whole Foods.xlsx",
    "Pur Gum Sprouts.xlsx": "Pur Gum Sprouts.xlsx",
    "Pur Gum The Fresh Market.xlsx": "Pur Gum The Fresh Market.xlsx",
    "Pur Gum Whole Foods.xlsx": "Pur Gum Whole Foods.xlsx",
    "Solely HEB.xlsx": "Solely HEB.xlsx",
    "Bachan Stop & Shop.xlsx": "Bachan Stop & Shop.xlsx",
    "Bachan Wakefern.xlsx": "Bachan Wakefern.xlsx",
    "Bachan Wegmans.xlsx": "Bachan Wegmans.xlsx",
    "ECOS NGVC.xlsx": "ECOS NGVC.xlsx",
    "ECOS WF.xlsx": "ECOS WF.xlsx",
    "Graza WF.xlsx": "Graza WF.xlsx",
    "GT Whole Foods.xlsx": "GT Whole Foods .xlsx",
    "Hain Independent Terra.xlsx": "Hain Independent Terra.xlsx",
    "Porta Whole Foods.xlsx": "Porta Whole Foods.xlsx",
    "Guayaki Yerba Mate Whole Foods.xlsx": "Guayaki Yerba Mate Whole Foods.xlsx",
    "Joseph's Pita.xlsx": "Joseph's Pita.xlsx",
    "Pop Secret- HEB (Our Home).xlsx": "Pop Secret- HEB (Our Home).xlsx",
    "Grazeful.xlsx": "Grazeful.xlsx",
    "Hain Independent Garden of Eatin.xlsx": "Hain Independent Garden of Eatin.xlsx",
    "Back 2 Nature The Fresh Market.xlsx": "Back 2 Nature The Fresh Market.xlsx",
}
def append_data(original_file, downloaded_file):
    try:
        # Cargar el archivo original
        wb_original = load_workbook(original_file)
        ws_original = wb_original.active
        
        # Cargar el archivo descargado
        wb_downloaded = load_workbook(downloaded_file)
        ws_downloaded = wb_downloaded.active

        # Leer los encabezados del archivo original
        original_headers = [cell.value for cell in ws_original[1]]

        # Leer los encabezados del archivo descargado
        downloaded_headers = [cell.value for cell in ws_downloaded[1]]
        
        # Encontrar los encabezados que coinciden
        matching_headers = [header for header in downloaded_headers if header in original_headers]
        non_matching_headers = [header for header in downloaded_headers if header not in original_headers]

        print(f"{len(matching_headers)} de {len(downloaded_headers)} encabezados coinciden.")
        if non_matching_headers:
            print(f"Encabezados que no coinciden: {non_matching_headers}")
        
        # Crear un diccionario para mapear los índices de las columnas
        header_index_map = {header: original_headers.index(header) + 1 for header in matching_headers}

        # Encontrar la primera fila vacía en el archivo original
        first_empty_row = ws_original.max_row + 1

        # Agregar los datos del archivo descargado al archivo original
        for row in ws_downloaded.iter_rows(min_row=2, values_only=True):
            new_row = [None] * len(original_headers)  # Crear una nueva fila vacía del tamaño de los encabezados originales
            for i, header in enumerate(downloaded_headers):
                if header in header_index_map:
                    new_row[header_index_map[header] - 1] = row[i]
            ws_original.append(new_row)

        # Guardar el archivo original con los datos agregados
        wb_original.save(original_file)
        print(f"Datos agregados de {downloaded_file} a {original_file}")
    except Exception as e:
        print(f"Error al procesar {downloaded_file}: {e}")

# Procesar cada archivo en el mapeo
for downloaded_file_name, original_file_name in file_mapping.items():
    downloaded_file_path = os.path.join(downloaded_files_directory, downloaded_file_name)
    original_file_path = os.path.join(original_files_directory, original_file_name)

    if os.path.exists(downloaded_file_path) and os.path.exists(original_file_path):
        append_data(original_file_path, downloaded_file_path)
    else:
        print(f"Archivo no encontrado: {downloaded_file_name} o {original_file_name}")
